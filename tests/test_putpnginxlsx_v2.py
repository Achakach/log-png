import pytest
import sys
import json
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import putpnginxlsx_v2


# --- _extract_device_from_filename tests ---

def test_device_extraction():
    assert (
        putpnginxlsx_v2._extract_device_from_filename(
            "HW-Core-BKK-01 display device.png"
        )
        == "HW-Core-BKK-01"
    )
    assert (
        putpnginxlsx_v2._extract_device_from_filename("HW-BR-Edge-01.png")
        == "HW-BR-Edge-01"
    )


# --- _group_pngs_by_device tests ---

def test_group_pngs_by_device():
    paths = [
        "scr/HW-C01 display device.png",
        "scr/HW-C01 display clock.png",
        "scr/HW-C02 display device.png",
        "scr/HW-C02 dis cur.png",
    ]
    groups = putpnginxlsx_v2._group_pngs_by_device(
        paths, ["display device", "display clock"]
    )
    assert list(groups.keys()) == ["HW-C01", "HW-C02"]
    assert len(groups["HW-C01"]) == 2
    assert len(groups["HW-C02"]) == 1
    assert "scr/HW-C02 dis cur.png" not in groups.get("HW-C02", [])



# --- _next_column tests ---

def test_next_column():
    assert putpnginxlsx_v2._next_column("B", 13) == "O"
    assert putpnginxlsx_v2._next_column("A", 3) == "D"


# --- _parse_cell tests ---

def test_parse_cell():
    assert putpnginxlsx_v2._parse_cell("B2") == ("B", 2)
    assert putpnginxlsx_v2._parse_cell("AA10") == ("AA", 10)


def test_parse_cell_invalid():
    with pytest.raises(ValueError):
        putpnginxlsx_v2._parse_cell("invalid")


# --- load_config tests ---

def test_load_config_validation(monkeypatch, tmp_path, capsys):
    cfg_file = tmp_path / "config.json"
    cfg_file.write_text(
        json.dumps(
            {
                "xlsx_input": "template.xlsx",
                "xlsx_output": "output.xlsx",
                "png_path": "screenshots/*.png",
                "sheet_configs": [["NetworkReport", "display device"]],
                "start_cell": "B2",
            }
        )
    )
    monkeypatch.setattr(
        putpnginxlsx_v2, "get_config_path", lambda: str(cfg_file)
    )

    with pytest.raises(SystemExit) as exc_info:
        putpnginxlsx_v2.load_config()

    assert exc_info.value.code == 1
    captured = capsys.readouterr()
    assert "Missing config fields" in captured.out


# --- main gallery layout tests ---

def test_main_gallery_layout(monkeypatch, tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    monkeypatch.setattr(putpnginxlsx_v2, "load_workbook", lambda path: wb)
    monkeypatch.setattr(putpnginxlsx_v2.os.path, "exists", lambda path: True)
    monkeypatch.setattr(wb, "save", lambda filename: None)

    cfg = {
        "xlsx_input": "fake.xlsx",
        "xlsx_output": str(tmp_path / "output.xlsx"),
        "png_path": "scr/*.png",
        "sheet_configs": [
            ["NetworkReport", "display device"],
            ["NetworkReport", "display clock"],
        ],
        "start_cell": "B2",
        "image_col_gap": 3,
        "device_row_gap": 3,
        "column_width_offset": 0,
        "row_height_offset": 0,
    }
    cfg_file = tmp_path / "config.json"
    cfg_file.write_text(json.dumps(cfg))

    monkeypatch.setattr(
        putpnginxlsx_v2, "get_config_path", lambda: str(cfg_file)
    )

    mock_pngs = [
        "scr/HW-C01 display device.png",
        "scr/HW-C01 display clock.png",
        "scr/HW-C02 display device.png",
    ]

    class MockGlob:
        @staticmethod
        def glob(pattern):
            return mock_pngs

    monkeypatch.setattr(putpnginxlsx_v2, "glob", MockGlob())

    class MockImg:
        def __init__(self, path):
            self.path = path
            self.width = 800
            self.height = 600

    monkeypatch.setattr(putpnginxlsx_v2, "OpenpyxlImage", MockImg)

    putpnginxlsx_v2.main()

    ws = wb["NetworkReport"]
    # Header row assertions
    assert ws["B1"].value == "display device"
    assert ws["B1"].font.bold
    assert ws["B1"].alignment.horizontal == "center"
    assert ws["B1"].alignment.vertical == "center"
    assert ws["F1"].value == "display clock"
    assert ws["F1"].font.bold
    assert ws["F1"].alignment.horizontal == "center"
    assert ws["F1"].alignment.vertical == "center"
    assert ws["A2"].value == "HW-C01"
    assert ws["A2"].font.bold
    assert ws["A2"].alignment.vertical == "center"
    assert ws["A2"].alignment.horizontal == "center"
    assert ws["A6"].value == "HW-C02"
    assert ws["A6"].font.bold
    assert ws["A6"].alignment.vertical == "center"
    assert ws["A6"].alignment.horizontal == "center"
    assert ws.row_dimensions[2].height == 450
    assert ws.row_dimensions[6].height == 450
    assert len(ws._images) == 3
    assert ws.column_dimensions["B"].width == 800 / 7


def test_column_width_offset(monkeypatch, tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    monkeypatch.setattr(putpnginxlsx_v2, "load_workbook", lambda path: wb)
    monkeypatch.setattr(putpnginxlsx_v2.os.path, "exists", lambda path: True)
    monkeypatch.setattr(wb, "save", lambda filename: None)

    cfg = {
        "xlsx_input": "fake.xlsx",
        "xlsx_output": str(tmp_path / "output.xlsx"),
        "png_path": "scr/*.png",
        "sheet_configs": [["NetworkReport", "display device"]],
        "start_cell": "B2",
        "image_col_gap": 3,
        "device_row_gap": 3,
        "column_width_offset": 5,
        "row_height_offset": 0,
    }
    cfg_file = tmp_path / "config.json"
    cfg_file.write_text(json.dumps(cfg))
    monkeypatch.setattr(putpnginxlsx_v2, "get_config_path", lambda: str(cfg_file))

    mock_pngs = ["scr/HW-C01 display device.png"]

    class MockGlob:
        @staticmethod
        def glob(pattern):
            return mock_pngs

    monkeypatch.setattr(putpnginxlsx_v2, "glob", MockGlob())

    class MockImg:
        def __init__(self, path):
            self.path = path
            self.width = 800
            self.height = 600

    monkeypatch.setattr(putpnginxlsx_v2, "OpenpyxlImage", MockImg)

    putpnginxlsx_v2.main()

    ws = wb["NetworkReport"]
    # Header row assertion
    assert ws["B1"].value == "display device"
    assert ws["B1"].font.bold
    assert ws["B1"].alignment.horizontal == "center"
    assert ws["B1"].alignment.vertical == "center"
    assert ws.column_dimensions["B"].width == 800 / 7 + 5
    assert len(ws._images) == 1


def test_row_height_offset(monkeypatch, tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    monkeypatch.setattr(putpnginxlsx_v2, "load_workbook", lambda path: wb)
    monkeypatch.setattr(putpnginxlsx_v2.os.path, "exists", lambda path: True)
    monkeypatch.setattr(wb, "save", lambda filename: None)

    cfg = {
        "xlsx_input": "fake.xlsx",
        "xlsx_output": str(tmp_path / "output.xlsx"),
        "png_path": "scr/*.png",
        "sheet_configs": [["NetworkReport", "display device"]],
        "start_cell": "B2",
        "image_col_gap": 3,
        "device_row_gap": 3,
        "column_width_offset": 0,
        "row_height_offset": 10,
    }
    cfg_file = tmp_path / "config.json"
    cfg_file.write_text(json.dumps(cfg))
    monkeypatch.setattr(putpnginxlsx_v2, "get_config_path", lambda: str(cfg_file))

    mock_pngs = ["scr/HW-C01 display device.png"]

    class MockGlob:
        @staticmethod
        def glob(pattern):
            return mock_pngs

    monkeypatch.setattr(putpnginxlsx_v2, "glob", MockGlob())

    class MockImg:
        def __init__(self, path):
            self.path = path
            self.width = 800
            self.height = 600

    monkeypatch.setattr(putpnginxlsx_v2, "OpenpyxlImage", MockImg)

    putpnginxlsx_v2.main()

    ws = wb["NetworkReport"]
    # Header row assertion
    assert ws["B1"].value == "display device"
    assert ws["B1"].font.bold
    assert ws["B1"].alignment.horizontal == "center"
    assert ws["B1"].alignment.vertical == "center"
    assert ws.row_dimensions[2].height == 600 * 0.75 + 10 * 5
    assert len(ws._images) == 1


def test_device_label_row_height_and_centered(monkeypatch, tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    monkeypatch.setattr(putpnginxlsx_v2, "load_workbook", lambda path: wb)
    monkeypatch.setattr(putpnginxlsx_v2.os.path, "exists", lambda path: True)
    monkeypatch.setattr(wb, "save", lambda filename: None)

    cfg = {
        "xlsx_input": "fake.xlsx",
        "xlsx_output": str(tmp_path / "output.xlsx"),
        "png_path": "scr/*.png",
        "sheet_configs": [["NetworkReport", "display device"]],
        "start_cell": "B2",
        "image_col_gap": 3,
        "device_row_gap": 3,
        "column_width_offset": 0,
        "row_height_offset": 0,
    }
    cfg_file = tmp_path / "config.json"
    cfg_file.write_text(json.dumps(cfg))
    monkeypatch.setattr(putpnginxlsx_v2, "get_config_path", lambda: str(cfg_file))

    mock_pngs = ["scr/HW-C01 display device.png"]

    class MockGlob:
        @staticmethod
        def glob(pattern):
            return mock_pngs

    monkeypatch.setattr(putpnginxlsx_v2, "glob", MockGlob())

    class MockImg:
        def __init__(self, path):
            self.path = path
            self.width = 800
            self.height = 600

    monkeypatch.setattr(putpnginxlsx_v2, "OpenpyxlImage", MockImg)

    putpnginxlsx_v2.main()

    ws = wb["NetworkReport"]
    # Header row assertion
    assert ws["B1"].value == "display device"
    assert ws["B1"].font.bold
    assert ws["B1"].alignment.horizontal == "center"
    assert ws["B1"].alignment.vertical == "center"
    assert ws["A2"].value == "HW-C01"
    assert ws["A2"].font.bold
    assert ws["A2"].alignment.horizontal == "center"
    assert ws["A2"].alignment.vertical == "center"
    assert ws.row_dimensions[2].height == 450
    assert len(ws._images) == 1
    assert ws.column_dimensions["B"].width == 800 / 7


def test_header_with_abbreviation(monkeypatch, tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    monkeypatch.setattr(putpnginxlsx_v2, "load_workbook", lambda path: wb)
    monkeypatch.setattr(putpnginxlsx_v2.os.path, "exists", lambda path: True)
    monkeypatch.setattr(wb, "save", lambda filename: None)

    cfg = {
        "xlsx_input": "fake.xlsx",
        "xlsx_output": str(tmp_path / "output.xlsx"),
        "png_path": "scr/*.png",
        "sheet_configs": [["NetworkReport", "dis cu"]],
        "start_cell": "B2",
        "image_col_gap": 3,
        "device_row_gap": 3,
        "column_width_offset": 0,
        "row_height_offset": 0,
    }
    cfg_file = tmp_path / "config.json"
    cfg_file.write_text(json.dumps(cfg))
    monkeypatch.setattr(putpnginxlsx_v2, "get_config_path", lambda: str(cfg_file))

    mock_pngs = ["scr/HW-C01 display current-configuration.png"]

    class MockGlob:
        @staticmethod
        def glob(pattern):
            return mock_pngs

    monkeypatch.setattr(putpnginxlsx_v2, "glob", MockGlob())

    class MockImg:
        def __init__(self, path):
            self.path = path
            self.width = 800
            self.height = 600

    monkeypatch.setattr(putpnginxlsx_v2, "OpenpyxlImage", MockImg)

    monkeypatch.setattr(
        putpnginxlsx_v2,
        "_load_abbreviations",
        lambda: [
            ["dis th", "display this"],
            ["dis cur", "display current-configuration"],
            ["dis cu", "display current-configuration"],
            ["dis", "display"],
            ["system", "system-view"],
            ["sys", "system-view"],
            ["comm", "commit"],
            ["q", "quit"],
            ["run", "run"],
        ],
    )

    putpnginxlsx_v2.main()

    ws = wb["NetworkReport"]
    assert ws["B1"].value == "display current-configuration"
    assert ws["B1"].font.bold
    assert ws["B1"].alignment.horizontal == "center"
    assert ws["B1"].alignment.vertical == "center"
    assert len(ws._images) == 1


def test_bidirectional_abbreviation_match(monkeypatch):
    """Config with full command matches PNG with abbreviation, and vice versa."""
    from putpnginxlsx_v2 import _group_pngs_by_device

    # Use known abbreviations directly (tuple list format)
    abbrev_list = [
        ("dis dev", "display device"),
        ("dis", "display"),
        ("dev", "device"),
    ]

    # Case 1: Config has full command, PNG has abbreviation
    mock_pngs = ["scr/HW-C01 dis dev.png"]
    groups = _group_pngs_by_device(mock_pngs, ["display device"], abbrev_list)
    assert "HW-C01" in groups

    # Case 2: Config has abbreviation, PNG has full command
    mock_pngs = ["scr/HW-C01 display device.png"]
    groups = _group_pngs_by_device(mock_pngs, ["dis dev"], abbrev_list)
    assert "HW-C01" in groups

    # Case 3: Both have abbreviations
    mock_pngs = ["scr/HW-C01 dis dev.png"]
    groups = _group_pngs_by_device(mock_pngs, ["dis dev"], abbrev_list)
    assert "HW-C01" in groups

    # Case 4: Both have full command
    mock_pngs = ["scr/HW-C01 display device.png"]
    groups = _group_pngs_by_device(mock_pngs, ["display device"], abbrev_list)
    assert "HW-C01" in groups
