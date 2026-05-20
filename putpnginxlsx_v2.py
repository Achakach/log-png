"""
putpnginxlsx_v2.py
==================
Utility script that inserts PNG screenshots into an Excel sheet
using a gallery layout: one row per device, images placed horizontally.
Auto-calculates image column span from PNG native pixel width.
Configuration is read from `putpnginxlsx_v2_config.json`.
If the config file is missing, a template will be auto-generated.
"""

import glob
import json

import os
import sys
import re

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import column_index_from_string, get_column_letter


CONFIG_FILENAME = "putpnginxlsx_v2_config.json"

_DEFAULT_CONFIG = {
    "xlsx_input": "template.xlsx",
    "xlsx_output": "output.xlsx",
    "png_path": "screenshots/*.png",
    "sheet_configs": [
        ["NetworkReport", "display device"],
        ["NetworkReport", "display clock"]
    ],
    "start_cell": "B2",
    "image_col_gap": 3,
    "device_row_gap": 3,
    "column_width_offset": 0,
    "row_height_offset": 0
}


def get_base_dir():
    """Get the directory where the .exe or .py script lives."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def get_config_path():
    return os.path.join(get_base_dir(), CONFIG_FILENAME)


def _ensure_config():
    """If config.json is missing, create a template and exit."""
    cfg_path = get_config_path()
    if os.path.exists(cfg_path):
        return
    with open(cfg_path, "w", encoding="utf-8") as f:
        json.dump(_DEFAULT_CONFIG, f, indent=2, ensure_ascii=False)
    print(f"Created template: {cfg_path}")
    print("Please edit it with your paths and keywords, then run again.")
    sys.exit(0)


def load_config():
    """Read and validate config.json."""
    _ensure_config()
    cfg_path = get_config_path()
    with open(cfg_path, "r", encoding="utf-8-sig") as f:
        cfg = json.load(f)

    raw_sheets = cfg.get("sheet_configs", [])
    cfg["sheet_configs"] = [
        tuple(item) for item in raw_sheets
        if isinstance(item, (list, tuple)) and len(item) == 2
    ]

    if not cfg["sheet_configs"]:
        print("ERROR: 'sheet_configs' must be a list of [sheet_name, keyword] pairs.")
        sys.exit(1)

    required = {
        "xlsx_input", "xlsx_output", "png_path", "sheet_configs",
        "start_cell", "image_col_gap", "device_row_gap",
        "column_width_offset",
        "row_height_offset"
    }
    missing = required - set(cfg.keys())
    if missing:
        print(f"ERROR: Missing config fields: {', '.join(sorted(missing))}")
        sys.exit(1)

    return cfg


def _parse_cell(cell_ref: str):
    """Parse a cell reference like 'B2' into (column_letter, row_number)."""
    match = re.match(r"^([A-Za-z]+)(\d+)$", cell_ref)
    if not match:
        raise ValueError(
            f"Invalid start_cell format: {cell_ref!r}. Expected format like 'B2'."
        )
    col = match.group(1).upper()
    row = int(match.group(2))
    return col, row


def _extract_device_from_filename(png_path):
    """Extract the device name (first token) from a PNG filename."""
    base = os.path.basename(png_path)
    name = base[:-4] if base.lower().endswith(".png") else base
    return name.split()[0] if name.split() else base


def _group_pngs_by_device(png_paths, keywords):
    """Filter PNGs by keywords and group by device name."""
    keywords_lower = [kw.lower() for kw in keywords]
    matched = []
    for p in png_paths:
        base_lower = os.path.basename(p).lower()
        if any(kw in base_lower for kw in keywords_lower):
            matched.append(p)

    groups = {}
    for p in matched:
        device = _extract_device_from_filename(p)
        groups.setdefault(device, []).append(p)

    for device in groups:
        groups[device].sort(key=lambda p: os.path.basename(p).lower())

    return dict(sorted(groups.items()))


def _native_dimensions(img_path):
    """Return the native (width, height) of a PNG in pixels."""
    img = OpenpyxlImage(img_path)
    return img.width, img.height


def _next_column(col_letter, offset):
    """Advance column letter by offset.
    _next_column('B', 13+3) -> 'P'
    """
    col_num = column_index_from_string(col_letter)
    col_num += offset
    return get_column_letter(col_num)


def main():
    cfg = load_config()

    xlsx_input = cfg["xlsx_input"]
    xlsx_output = cfg["xlsx_output"]
    png_path = cfg["png_path"]
    sheet_configs = cfg["sheet_configs"]
    start_cell = cfg["start_cell"]
    image_col_gap = cfg["image_col_gap"]
    device_row_gap = cfg["device_row_gap"]

    if not os.path.exists(xlsx_input):
        print(f"ERROR: Input file not found: {xlsx_input}")
        sys.exit(1)

    png_files = glob.glob(png_path)
    if not png_files:
        print(f"ERROR: No PNG files found matching: {png_path}")
        sys.exit(1)

    wb = load_workbook(xlsx_input)
    start_col, start_row = _parse_cell(start_cell)

    # Collect unique sheet names in order of first appearance
    seen_sheets = []
    for sheet_name, _ in sheet_configs:
        if sheet_name not in seen_sheets:
            seen_sheets.append(sheet_name)

    for sheet_name in seen_sheets:
        # Gather keywords for this sheet, preserving order
        keywords = []
        seen_keywords = set()
        for sn, kw in sheet_configs:
            if sn == sheet_name and kw not in seen_keywords:
                keywords.append(kw)
                seen_keywords.add(kw)

        # Group PNGs by device
        device_groups = _group_pngs_by_device(png_files, keywords)

        if not device_groups:
            print(
                f"WARNING: No PNGs matched keywords for sheet '{sheet_name}'. Skipping."
            )
            continue

        # Get or create sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            print(f"INFO: Created new sheet '{sheet_name}'.")

        current_row = start_row

        for device, device_pngs in device_groups.items():
            # Write device name in column A
            cell = ws[f"A{current_row}"]
            cell.value = device
            cell.font = Font(bold=True)

            current_col = start_col
            max_target_h_pixels = 0

            for keyword in keywords:
                keyword_lower = keyword.lower()
                matched_png = None
                for p in device_pngs:
                    if keyword_lower in os.path.basename(p).lower():
                        matched_png = p
                        break

                if matched_png:
                    try:
                        native_w, native_h = _native_dimensions(matched_png)
                    except Exception as e:
                        print(
                            f"WARNING: Could not read image dimensions for "
                            f"{os.path.basename(matched_png)}: {e}"
                        )
                        native_w = 192
                        native_h = 144

                    target_w_pixels = native_w
                    scale = target_w_pixels / native_w if native_w else 1
                    target_h_pixels = int(native_h * scale) if native_w else native_h
                    max_target_h_pixels = max(max_target_h_pixels, target_h_pixels)

                    img = OpenpyxlImage(matched_png)
                    img.width = target_w_pixels
                    img.height = target_h_pixels

                    cell_ref = f"{current_col}{current_row}"
                    ws.add_image(img, cell_ref)
                    ws.column_dimensions[current_col].width = max(
                        12, target_w_pixels / 7 + cfg["column_width_offset"]
                    )
                    print(
                        f"[{sheet_name}] {device} / "
                        f"{os.path.basename(matched_png)} at {cell_ref} "
                        f"({img.width}x{img.height}px)"
                    )

                    current_col = _next_column(current_col, 1 + image_col_gap)
                else:
                    # No image for this keyword; advance by a default span
                    current_col = _next_column(current_col, 1 + image_col_gap)

            if max_target_h_pixels > 0:
                ws.row_dimensions[current_row].height = max_target_h_pixels * 0.75 + cfg["row_height_offset"]
            ws[f"A{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

            current_row += 1 + device_row_gap

        # Auto-fit column A width based on longest device name
        if device_groups:
            longest = max(len(d) for d in device_groups)
            ws.column_dimensions["A"].width = min(50, max(12, longest + 2))

    wb.save(xlsx_output)
    print(f"Saved: {xlsx_output}")


if __name__ == "__main__":
    main()
