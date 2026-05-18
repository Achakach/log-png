import pytest
import sys
import json
import math
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import putpnginxlsx_v3


# --- _extract_device_from_filename tests ---

def test_device_extraction():
    assert (
        putpnginxlsx_v3._extract_device_from_filename(
            "HW-Core-BKK-01 display device.png"
        )
        == "HW-Core-BKK-01"
    )
    assert (
        putpnginxlsx_v3._extract_device_from_filename("HW-BR-Edge-01.png")
        == "HW-BR-Edge-01"
    )


# --- _group_pngs_by_device tests ---

def test_group_pngs_by_device():
    paths = [
        "scr/HW-C01 display device.png",
        "scr/HW-C01 display clock.png",
        "scr/HW-C02 display device.png",
        "scr/HW-C02 dis cur.png",
    ]
    groups = putpnginxlsx_v3._group_pngs_by_device(
        paths, ["display device", "display clock"]
    )
    assert list(groups.keys()) == ["HW-C01", "HW-C02"]
    assert len(groups["HW-C01"]) == 2
    assert len(groups["HW-C02"]) == 1
    assert "scr/HW-C02 dis cur.png" not in groups.get("HW-C02", [])


# --- _rows_for_image tests ---

def test_rows_for_image(monkeypatch):
    class MockImage:
        def __init__(self, path):
            self.width = 1200
            self.height = 1260

    monkeypatch.setattr(putpnginxlsx_v3, "OpenpyxlImage", MockImage)
    result = putpnginxlsx_v3._rows_for_image("fake.png")
    assert result == (63, 1200, 1260)

    class MockImageSmall:
        def __init__(self, path):
            self.width = 1200
            self.height = 100

    monkeypatch.setattr(putpnginxlsx_v3, "OpenpyxlImage", MockImageSmall)
    result = putpnginxlsx_v3._rows_for_image("fake.png")
    assert result == (5, 1200, 100)


# --- _parse_cell tests ---

def test_parse_cell():
    assert putpnginxlsx_v3._parse_cell("B2") == ("B", 2)
    assert putpnginxlsx_v3._parse_cell("AA10") == ("AA", 10)


def test_parse_cell_invalid():
    with pytest.raises(ValueError):
        putpnginxlsx_v3._parse_cell("invalid")


# --- _next_column tests ---

def test_next_column():
    assert putpnginxlsx_v3._next_column("B", 1) == "C"
    assert putpnginxlsx_v3._next_column("A", 3) == "D"


# --- main vertical layout tests ---

def test_main_vertical_layout(monkeypatch, tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    monkeypatch.setattr(putpnginxlsx_v3, "load_workbook", lambda path: wb)
    monkeypatch.setattr(putpnginxlsx_v3.os.path, "exists", lambda path: True)
    monkeypatch.setattr(wb, "save", lambda filename: None)

    cfg = {
        "xlsx_input": "template.xlsx",
        "xlsx_output": str(tmp_path / "output.xlsx"),
        "png_path": "scr/*.png",
        "sheet_configs": [
            ["NetworkReport", "display device"],
            ["NetworkReport", "display clock"],
        ],
        "start_cell": "B2",
        "image_col_gap": 65,
        "device_row_gap": 3,
    }
    cfg_file = tmp_path / "config.json"
    cfg_file.write_text(json.dumps(cfg))

    monkeypatch.setattr(
        putpnginxlsx_v3, "get_config_path", lambda: str(cfg_file)
    )

    mock_pngs = [
        "scr/HW-C01 display device.png",
        "scr/HW-C01 display clock.png",
        "scr/HW-C02 display device.png",
    ]

    class MockGlob:
        @staticmethod
        def glob(pattern):
            if "*.png" in pattern:
                return mock_pngs
            return []

    monkeypatch.setattr(putpnginxlsx_v3, "glob", MockGlob())

    class MockImg:
        def __init__(self, path):
            self.width = 1200
            self.height = 1260

    monkeypatch.setattr(putpnginxlsx_v3, "OpenpyxlImage", MockImg)

    putpnginxlsx_v3.main()

    ws = wb["NetworkReport"]
    # First keyword label
    assert ws["A2"].value == "display device"
    assert ws["A2"].font.bold
    # Second keyword label at row 2 + 63 (row_span) + 65 (image_col_gap) = 130
    assert ws["A130"].value == "display clock"
    assert ws["A130"].font.bold
    # Total images placed
    assert len(ws._images) == 3
    # Column widths
    assert ws.column_dimensions["B"].width == 18.75
    assert ws.column_dimensions["A"].width == 16
