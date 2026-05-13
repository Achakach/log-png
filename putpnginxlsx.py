"""
putpnginxlsx.py
===============
Utility script that filters PNG screenshots by command keyword and inserts
them vertically into a specified Excel sheet.

Dependencies
------------
    pip install openpyxl

Usage
-----
    python putpnginxlsx.py

Configuration is read from `config.json` in the same directory.
If `config.json` is missing, a template will be auto-generated.
"""

import glob
import json
import math
import os
import sys
import re

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage


CONFIG_FILENAME = "putpnginxlsx_config.json"


_DEFAULT_CONFIG = {
    "xlsx_input": "template.xlsx",
    "xlsx_output": "output.xlsx",
    "png_path": "screenshots/*.png",
    "sheet_configs": [
        ["DisplayClock", "display clock"],
        ["DisplayDevice", "display device"]
    ],
    "start_cell": "B2",
    "image_width_inches": 6,
    "gap_rows": 3
}


# --- Excel dimensional constants ---
_ROW_HEIGHT_INCHES = 0.21           # Default Excel row height ≈ 15pt


def _ensure_config():
    """If config.json is missing, create a template and exit."""
    if os.path.exists(CONFIG_FILENAME):
        return
    with open(CONFIG_FILENAME, "w", encoding="utf-8") as f:
        json.dump(_DEFAULT_CONFIG, f, indent=2, ensure_ascii=False)
    print(f"Created template: {CONFIG_FILENAME}")
    print("Please edit it with your paths and keywords, then run again.")
    sys.exit(0)


def load_config():
    """Read and validate config.json."""
    _ensure_config()
    with open(CONFIG_FILENAME, "r", encoding="utf-8") as f:
        cfg = json.load(f)

    # Convert sheet_configs from list-of-lists to list-of-tuples
    raw_sheets = cfg.get("sheet_configs", [])
    cfg["sheet_configs"] = [tuple(item) for item in raw_sheets if isinstance(item, (list, tuple)) and len(item) == 2]

    if not cfg["sheet_configs"]:
        print("ERROR: 'sheet_configs' must be a list of [sheet_name, keyword] pairs.")
        sys.exit(1)

    return cfg


def _parse_cell(cell_ref: str):
    """Parse a cell reference like 'B2' into (column_letter, row_number)."""
    match = re.match(r"^([A-Za-z]+)(\d+)$", cell_ref)
    if not match:
        raise ValueError(f"Invalid start_cell format: {cell_ref!r}. Expected format like 'B2'.")
    col = match.group(1).upper()
    row = int(match.group(2))
    return col, row


def main():
    cfg = load_config()

    xlsx_input = cfg["xlsx_input"]
    xlsx_output = cfg["xlsx_output"]
    png_path = cfg["png_path"]
    sheet_configs = cfg["sheet_configs"]
    start_cell = cfg["start_cell"]
    image_width_inches = cfg["image_width_inches"]
    gap_rows = cfg["gap_rows"]

    # Validate input file
    if not os.path.exists(xlsx_input):
        print(f"ERROR: Input file not found: {xlsx_input}")
        sys.exit(1)

    # Parse PNG glob once
    png_files = glob.glob(png_path)
    if not png_files:
        print(f"ERROR: No PNG files found matching: {png_path}")
        sys.exit(1)

    # Load workbook once
    wb = load_workbook(xlsx_input)
    target_width_px = image_width_inches * 96  # Excel uses 96 DPI

    # Track next available row per sheet so repeated sheet names append vertically
    sheet_next_row = {}

    for sheet_name, keyword in sheet_configs:
        keyword_lower = keyword.lower()
        filtered = [p for p in png_files if keyword_lower in os.path.basename(p).lower()]

        if not filtered:
            print(f"WARNING: No PNGs matched keyword '{keyword}' for sheet '{sheet_name}'. Skipping.")
            continue

        # Sort by device name (first token of filename)
        def sort_key(path):
            base = os.path.basename(path)
            name = base[:-4] if base.lower().endswith(".png") else base
            return name.split()[0] if name.split() else base
        filtered.sort(key=sort_key)

        # Determine start row for this sheet
        if sheet_name in sheet_next_row:
            start_col, current_row = _parse_cell(start_cell)
            current_row = sheet_next_row[sheet_name]
        else:
            start_col, current_row = _parse_cell(start_cell)

        # Get or create sheet
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            print(f"INFO: Created new sheet '{sheet_name}'.")

        for png_path in filtered:
            cell_ref = f"{start_col}{current_row}"

            img = OpenpyxlImage(png_path)
            native_w, native_h = img.width, img.height
            scale = target_width_px / native_w
            img.width = target_width_px
            img.height = int(native_h * scale)

            img_height_inches = img.height / 96
            rows_occupied = math.ceil(img_height_inches / _ROW_HEIGHT_INCHES)

            ws.add_image(img, cell_ref)
            print(f"[{sheet_name}] Inserted: {os.path.basename(png_path)} at {cell_ref} ({img.width}x{img.height}px)")

            current_row += rows_occupied + gap_rows

        # Save the next available row for this sheet (for subsequent keyword batches)
        sheet_next_row[sheet_name] = current_row

    # Save output (never overwrite input directly)
    wb.save(xlsx_output)
    print(f"Saved: {xlsx_output}")


if __name__ == "__main__":
    main()
